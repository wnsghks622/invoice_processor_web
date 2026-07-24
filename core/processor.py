# -*- coding: utf-8 -*-
"""
Invoice Processor - Folder + Claude AI + Excel
Compatible with Windows 10/11, Python 3.8+

Scans a local folder for invoice files (PDFs or images), extracts invoice data from each
using Claude AI, writes the results to a property-tabbed Excel file (plus an "All Invoices"
log), and files every processed PDF into processed/<property>/. It also writes a per-property
_amounts.csv sidecar that the Bank Rec assembler (bankrec.py) reads, so the two tools share
verified amounts. See README.md for the full month-end bank-rec workflow.
"""

from __future__ import annotations

import os
import sys
import base64
import csv
import difflib
import json
import re
import shutil
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict

# Load .env file if present (so users don't have to set system env vars)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # dotenv optional; env vars can be set manually


# - CONFIG -
#
# This is the DB-backed (SQLite) build of the processor. Data lives in the web app's data/
# folder (config.py); properties and vendors come from the database, not CSV files. The
# original CSV/xlsx-based processor is kept, untouched, in the sibling invoice_processor/ folder.
import config                       # web app paths (data/ dir, DB path)
from . import db                    # SQLite data layer (cycle-safe: only used at call time)

BASE_DIR = config.HERE

# Folder you drop invoices into, and where files are filed once processed - the app's own data/.
INPUT_FOLDER     = config.INVOICES_TO_PROCESS
PROCESSED_FOLDER = config.PROCESSED

# Retained only so legacy references resolve; the DB is the source of truth now (nothing is
# written to this path).
EXCEL_OUTPUT     = config.DATA_DIR / "invoices.xlsx"

# Property/vendor lists now come from the database (see load_property_list / load_vendor_map,
# which ignore these paths). Kept as harmless defaults for any old call site.
PROPERTY_LIST_FILE = config.DATA_DIR / "properties.csv"
VENDOR_MAP_FILE    = config.DATA_DIR / "vendors.csv"

# Claude model used for extraction. Override in .env (e.g. CLAUDE_MODEL=claude-opus-4-5)
# without touching the code. Defaults to Haiku 4.5 - the cheapest; step up to a larger
# model for higher accuracy on messy or handwritten invoices.
CLAUDE_MODEL = os.environ.get("CLAUDE_MODEL", "").strip() or "claude-haiku-4-5"


# - WINDOWS CONSOLE FIX -
# Force UTF-8 output so emoji/unicode don't crash Windows CMD/PowerShell

def setup_console():
    if sys.platform == "win32":
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")

setup_console()


def _interactive() -> bool:
    """True only when a real terminal is attached. False when the web app runs us headless via
    subprocess, so we never block on input()."""
    try:
        return bool(sys.stdin and sys.stdin.isatty())
    except Exception:
        return False


def _pause(msg: str = "Press Enter to close...") -> None:
    """input() that no-ops when there's no terminal (headless web run)."""
    if _interactive():
        try:
            input(msg)
        except EOFError:
            pass


# - CLAUDE AI EXTRACTION -

# Tokens (__LIKE_THIS__) are filled in by build_extraction_prompt via str.replace, so the
# JSON braces below stay as plain, readable single braces (no .format escaping needed).
_PROMPT_TEMPLATE = """You are an invoice data extraction assistant.
A single document may contain ONE invoice or SEVERAL invoices/bills - for example one bill
per unit, apartment, suite, or account at the same property. Extract EVERY invoice you find.

Return ONLY a valid JSON object of exactly this shape - no explanation, no markdown fences:

{
  "invoices": [
    {
      "vendor_name": "...",
      "invoice_number": "the invoice / bill / statement number, or null (NEVER an account number)",
      "unit": "the unit / apartment / suite number this bill is for, or null",
      "invoice_date": "MM/DD/YYYY or as written, or null",
      "due_date": "MM/DD/YYYY or as written, or null",
      "total_amount": "the total amount due / balance due, with currency symbol, or null",
      "description": "one short sentence describing what the invoice is for",
      "line_items": [{"description": "what this charge is for", "amount": "this line's amount"}],
      "property": __PROPERTY_FIELD__
    }
  ]
}

Return one array element per invoice/bill. If the document is a single invoice, return an
array with exactly one element.
If a field is not found, use null.
For vendor_name, use the company or person sending the invoice (the biller).
For invoice_number, use ONLY a genuine invoice / bill / statement number. Do NOT use an
account number, customer number, meter number, or service ID. Utility bills (water, power,
gas, etc.) often show only an account number and no invoice number - in that case return
null for invoice_number (a date-based number will be filled in automatically).
For unit, capture the apartment / unit / suite number (e.g. "APT 205", "Suite 3", "Unit B")
when the bill is for a specific unit; otherwise null. Bills for different units of the SAME
property must each be their own separate array element - do not merge them.
But do NOT split ONE invoice into several elements: if a single invoice number covers several
units or line items under one grand total, that is ONE invoice - return ONE element and put
those charges in line_items. Only return separate elements for physically separate bills, each
with its OWN invoice number or its OWN total. Never repeat the same invoice number AND the same
total on more than one element.
For total_amount, use the invoice's final TOTAL or BALANCE DUE (the amount actually owed) -
not a subtotal, a tax line, or a single line-item price.
For line_items, list every separate charge/service line with its description and amount.
If the invoice is a single charge with no itemized breakdown, return an empty list [].
__PROPERTY_GUIDANCE__"""

_PROPERTY_GUIDANCE_WITH_LIST = """
PROPERTY = THE LOCATION WHERE THE SERVICE/WORK WAS PERFORMED.
Find the service location on the invoice - usually labeled "Location", "Service Location",
"Service Address", "Job Site", "Site", or "Property". Then match THAT location to exactly
one entry in the KNOWN PROPERTIES list below, using its name or any of its addresses/
aliases, and return that entry's canonical name EXACTLY as written in the list.

Never use any of these as the property:
  - the "Bill To", "Remit To", or payment/mailing address
  - the property-management company's office address
  - the vendor's own address, letterhead, or "make check payable to" address

If you cannot confidently match the service location to a property in the list, return
null for "property". Do not guess, and never return a property that is not in the list.

KNOWN PROPERTIES (canonical name -- addresses/aliases):
__PROPERTY_REFERENCE__
"""

_PROPERTY_GUIDANCE_NO_LIST = """
For property, use the LOCATION WHERE THE SERVICE/WORK WAS PERFORMED - look for "Location",
"Service Location", "Service Address", "Job Site", or "Site". Do NOT use the "Bill To" /
remit-to address, the property-management company office, or the vendor's own address.
"""


def build_extraction_prompt(property_reference: str = "") -> str:
    """
    Build the extraction prompt. When a property list is available it's embedded so Claude
    returns canonical properties (the service locations), never billing addresses. The prompt
    always asks for an "invoices" array, so multi-unit / multi-bill PDFs are fully captured.
    """
    if property_reference:
        field = ('"the canonical property name from the KNOWN PROPERTIES list below - the '
                 'location where the service was performed - or null"')
        guidance = _PROPERTY_GUIDANCE_WITH_LIST.replace("__PROPERTY_REFERENCE__", property_reference)
    else:
        field = ('"name or address of the location where the service was performed '
                 '(NOT the billing address), or null"')
        guidance = _PROPERTY_GUIDANCE_NO_LIST
    return (_PROMPT_TEMPLATE
            .replace("__PROPERTY_FIELD__", field)
            .replace("__PROPERTY_GUIDANCE__", guidance))


def _coerce_invoice_list(parsed) -> list:
    """
    Normalize Claude's JSON into a list of invoice dicts. Accepts the new {"invoices": [...]}
    shape, a bare list, or a single invoice object (older shape) - always returns a list.
    """
    if isinstance(parsed, dict):
        items = parsed["invoices"] if isinstance(parsed.get("invoices"), list) else [parsed]
    elif isinstance(parsed, list):
        items = parsed
    else:
        return []
    return [it for it in items if isinstance(it, dict)]


def _heal_pdf_bytes(data: bytes) -> bytes:
    """Some vendor PDFs (notably SoCal Gas) start with a junk header such as
    '%%BILLID : ...' instead of '%PDF'. The Claude API rejects those as an invalid PDF even
    though the real document begins a little further in. Trim anything before the first
    '%PDF' marker so extraction succeeds. (The assembler heals the same way via _fixed_bytes.)"""
    if data[:4] != b"%PDF":
        i = data.find(b"%PDF")
        if i > 0:
            return data[i:]
    return data


def extract_invoice_data(
    client: anthropic.Anthropic,
    file_bytes: bytes,
    media_type: str,
    prompt: str,
) -> list:
    """Send a file to Claude and return a LIST of invoice dicts (a file may hold several)."""
    if media_type == "application/pdf":
        file_bytes = _heal_pdf_bytes(file_bytes)
    b64 = base64.standard_b64encode(file_bytes).decode("utf-8")

    if media_type == "application/pdf":
        content = [
            {
                "type": "document",
                "source": {
                    "type": "base64",
                    "media_type": "application/pdf",
                    "data": b64,
                },
            },
            {"type": "text", "text": prompt},
        ]
    else:
        content = [
            {
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": media_type,
                    "data": b64,
                },
            },
            {"type": "text", "text": prompt},
        ]

    try:
        message = client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=8000,   # room for many bills (one per unit) each with line items
            messages=[{"role": "user", "content": content}],
        )
        raw = message.content[0].text.strip()
        # Strip markdown fences if model adds them
        raw = re.sub(r"^```json\s*|```$", "", raw, flags=re.MULTILINE).strip()
        return _coerce_invoice_list(json.loads(raw))
    except anthropic.AuthenticationError:
        # A bad API key fails identically on every file - let main() stop the run
        # cleanly instead of repeating the same error for each invoice.
        raise
    except (json.JSONDecodeError, IndexError, anthropic.APIError) as exc:
        print(f"    [!] Claude extraction failed: {exc}")
        return []


# - EXCEL OUTPUT -

SUMMARY_SHEET = "Summary"
MASTER_SHEET  = "All Invoices"          # complete log of every invoice ever processed
RESERVED_SHEETS = {SUMMARY_SHEET, MASTER_SHEET}

# Master log = full audit trail. The Status column flags duplicates.
MASTER_HEADER = [
    "Status",
    "Vendor Name",
    "Invoice #",
    "Unit",
    "Invoice Date",
    "Due Date",
    "Amount",
    "Description",
    "Line Items (review)",
    "Property",
    "Source File",
    "Date Processed",
    "Entered in Yardi",
    "Stored File",        # the filed VendorShort_MM_YYYY.pdf name - join key for the amount sidecar
    "Reconciled",         # set by reconcile.py to "<Month YYYY>" once the item clears a statement
    "Check #",            # YOU type this after cutting the check; lets the assembler match the
                          # invoice to the cleared check on the bank statement (strongest signal)
    "Carried Forward",    # reconcile.py stamps the month an invoice was reviewed but did NOT clear;
                          # it stays pending and is watched against next month's statement
]
MASTER_WIDTHS = [12, 22, 15, 12, 12, 12, 12, 30, 36, 20, 22, 14, 16, 24, 18, 12, 16]

# Each property tab holds the first-time invoices for that property. Duplicates are not
# repeated here - they're flagged in the master log only - so property tabs stay clean.
PROPERTY_HEADER = [
    "Vendor Name",
    "Invoice #",
    "Unit",
    "Invoice Date",
    "Due Date",
    "Amount",
    "Description",
    "Line Items (review)",
    "Source File",
    "Date Processed",
    "Entered in Yardi",
    "Stored File",        # mirrors the master log (sidecar join key)
    "Reconciled",
    "Check #",
    "Carried Forward",
]
PROPERTY_WIDTHS = [24, 16, 12, 12, 12, 12, 34, 40, 26, 14, 16, 24, 18, 12, 16]

HEADER_FILL = PatternFill("solid", start_color="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ALT_FILL    = PatternFill("solid", start_color="EBF0FA")
DATA_FONT   = Font(name="Arial", size=10)
DUP_FILL    = PatternFill("solid", start_color="C00000")   # red - flags duplicate rows
DUP_FONT    = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BORDER_SIDE = Side(style="thin", color="C0C0C0")
THIN_BORDER = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE,  bottom=BORDER_SIDE,
)

# "Entered in Yardi" checkoff column: a one-click check-mark dropdown on every row that
# turns the cell green once chosen, for tracking which invoices you've keyed into Yardi.
YARDI_COL       = "Entered in Yardi"
YARDI_CHECK     = "✓"                                    # check mark
YARDI_DONE_FILL = PatternFill("solid", start_color="C6EFCE")  # green once checked
YARDI_DONE_FONT = Font(bold=True, color="006100", name="Arial", size=11)


def _sanitize_sheet_name(name: str) -> str:
    """Excel sheet names: max 31 chars, no special characters, never blank, and never
    colliding with our Summary / All Invoices sheets."""
    clean = re.sub(r'[:\\/?*\[\]]', "", str(name)).strip()
    clean = clean[:31].strip() or "Unknown"
    if clean in RESERVED_SHEETS:
        clean = f"{clean[:21]} (property)"
    return clean


def _style_header_row(ws, header) -> None:
    for col_idx, title in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
    ws.row_dimensions[1].height = 22


def _set_col_widths(ws, widths) -> None:
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _append_styled_row(ws, header, values, duplicate=False) -> None:
    """Append one styled data row. Numeric values (amounts) get currency formatting and
    right alignment. On duplicates, paint the Status cell (col 1) red."""
    row_num = ws.max_row + 1
    fill = ALT_FILL if row_num % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
    for col_idx in range(1, len(header) + 1):
        value = values[col_idx - 1] if col_idx - 1 < len(values) else ""
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font      = DATA_FONT
        cell.fill      = fill
        cell.border    = THIN_BORDER
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            cell.number_format = '$#,##0.00'
            cell.alignment     = Alignment(horizontal="right", vertical="center")
        elif header[col_idx - 1] == YARDI_COL:
            cell.alignment     = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment     = Alignment(vertical="center", wrap_text=True)
    if duplicate:
        cell = ws.cell(row=row_num, column=1)
        cell.fill      = DUP_FILL
        cell.font      = DUP_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row_num].height = 18


def _add_yardi_checkoff(ws, header) -> None:
    """
    Turn the 'Entered in Yardi' column into a click-to-check column: a dropdown whose only
    value is a check mark, plus a green highlight once it's chosen. Applied to the whole
    column so every current and future row has it. Typing is still allowed (no error popup).
    """
    if YARDI_COL not in header:
        return
    col = get_column_letter(header.index(YARDI_COL) + 1)
    rng = f"{col}2:{col}1048576"
    dv = DataValidation(type="list", formula1=f'"{YARDI_CHECK}"', allow_blank=True)
    dv.showErrorMessage = False
    ws.add_data_validation(dv)
    dv.add(rng)
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=[f'"{YARDI_CHECK}"'],
                   fill=YARDI_DONE_FILL, font=YARDI_DONE_FONT),
    )


def ensure_columns(ws, header, widths=None) -> None:
    """
    Append any header columns missing from an existing sheet (label + width), so a sheet
    created before a column was added still gains it. Rows are appended by position, so new
    columns must go at the END - this brings older sheets up to the current header without
    shifting their data. Name-based and idempotent (running it twice is a no-op).
    """
    have = [c.value for c in ws[1]]
    for name in header:
        if name in have:
            continue
        col = len(have) + 1
        c = ws.cell(row=1, column=col, value=name)
        c.font, c.fill, c.border = HEADER_FONT, HEADER_FILL, THIN_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if widths:
            ws.column_dimensions[get_column_letter(col)].width = widths[header.index(name)]
        have.append(name)


def _get_or_create_master_sheet(wb):
    if MASTER_SHEET in wb.sheetnames:
        return wb[MASTER_SHEET]
    idx = 1 if SUMMARY_SHEET in wb.sheetnames else 0
    ws = wb.create_sheet(title=MASTER_SHEET, index=idx)
    _style_header_row(ws, MASTER_HEADER)
    _set_col_widths(ws, MASTER_WIDTHS)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASTER_HEADER))}1"
    _add_yardi_checkoff(ws, MASTER_HEADER)
    return ws


def _get_or_create_property_sheet(wb, property_name: str):
    sheet_name = _sanitize_sheet_name(property_name or "Unknown Property")
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    _style_header_row(ws, PROPERTY_HEADER)
    _set_col_widths(ws, PROPERTY_WIDTHS)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(PROPERTY_HEADER))}1"
    _add_yardi_checkoff(ws, PROPERTY_HEADER)
    return ws


# Date formats Claude may return (prompt asks for MM/DD/YYYY "or as written")
_DATE_FORMATS = (
    "%m/%d/%Y", "%m-%d-%Y", "%m.%d.%Y",
    "%Y-%m-%d", "%Y/%m/%d",
    "%m/%d/%y", "%m-%d-%y",
    "%B %d, %Y", "%b %d, %Y",
    "%d %B %Y", "%d %b %Y",
    "%d-%b-%Y", "%d-%B-%Y",
)

# Values that mean "field not found" even when Claude returns text instead of null
_MISSING_VALUES = frozenset({
    "", "null", "none", "nil", "n/a", "na", "n.a.", "-", "--",
    "not available", "not found", "not provided", "unknown", "tbd",
})


def _is_missing(value: Optional[str]) -> bool:
    """True if a value is absent or a placeholder (null / N/A / unknown / all-zeros)."""
    if value is None:
        return True
    text = str(value).strip()
    if text.lower() in _MISSING_VALUES:
        return True
    # Treat all-zero numbers (0, 00, 000, 0-0, ...) as placeholders. Drop separators
    # first so "0-0" counts, but keep real numbers like "0012345" that have other digits.
    digits = re.sub(r"[^0-9A-Za-z]", "", text)
    return bool(digits) and set(digits) == {"0"}


def _parse_date(date_str: Optional[str]) -> Optional[datetime]:
    """Parse a date string into a datetime using the known formats, or None."""
    if _is_missing(date_str):
        return None
    raw = str(date_str).strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


def _format_date_mmddyyyy(date_str: Optional[str]) -> Optional[str]:
    """Parse a date string and return it as MMDDYYYY, or None if unparseable."""
    dt = _parse_date(date_str)
    return dt.strftime("%m%d%Y") if dt else None


def resolve_invoice_number(data: Dict) -> str:
    """
    Return the invoice number, or fall back to the invoice date (then due date)
    formatted as MMDDYYYY when the invoice has no number on it.
    """
    raw = data.get("invoice_number")
    if not _is_missing(raw):
        return str(raw).strip()

    # No invoice number found - substitute a date in MMDDYYYY format
    for date_field in ("invoice_date", "due_date"):
        formatted = _format_date_mmddyyyy(data.get(date_field))
        if formatted:
            return formatted

    return ""


def _is_date_based_number(invoice_no, invoice_date, due_date) -> bool:
    """
    True if `invoice_no` is one of our date-based fallbacks - the MMDDYYYY stand-in
    resolve_invoice_number() fills in when an invoice has no real number - rather than a
    genuine invoice number. Computed the same way from a live invoice and from a saved row,
    so the duplicate key below comes out identical whether it's built during a run or rebuilt
    from the spreadsheet on a later run.
    """
    digits = re.sub(r"\D", "", str(invoice_no or ""))
    if len(digits) != 8:                       # MMDDYYYY is always 8 digits
        return False
    return digits in {_format_date_mmddyyyy(invoice_date), _format_date_mmddyyyy(due_date)}


def _invoice_key(vendor: str, invoice_no: str, unit: str = "",
                 property_name: str = "", amount=None, *, date_based: bool = False) -> str:
    """
    Identity used to spot duplicates: normalized vendor + invoice number + unit + PROPERTY.
    Returns '' when there's no invoice number. The unit keeps separate bills for different
    units of one property apart; the property keeps the SAME vendor + invoice number issued for
    two DIFFERENT properties from being flagged as duplicates of each other.

    When the "number" is a date-based fallback (`date_based`), the date is NOT a unique id, so
    the amount is folded in too (two unrelated same-day bills from one vendor at one property are
    told apart by amount). The key layout is vendor|invoice#|unit|property[|amount] - so a
    genuine invoice number is a strong vendor+number+unit+property key.
    """
    inv = _normalize_key(invoice_no or "")
    if not inv:
        return ""
    parts = [_normalize_key(vendor or ""), inv, _normalize_key(unit or ""),
             _normalize_key(property_name or "")]
    if date_based:
        amt = _parse_amount(amount)
        parts.append(f"{amt:.2f}" if amt is not None else "")
    return "|".join(parts)


def merge_oversplit_invoices(invoices):
    """Fold back a single invoice the extractor split into one element per line-item/unit.
    Elements sharing a real (non-null) invoice number AND the same total are the same invoice
    (one number, one grand total) - keep one row and gather the distinct units into it, so the
    amount isn't counted once per unit. Genuinely separate bills (their own invoice number, or
    their own per-unit total) get different keys and are left untouched."""
    out, by_key = [], {}
    for data in invoices:
        inv = _normalize_key(str(data.get("invoice_number") or ""))
        amt = _parse_amount(data.get("total_amount"))
        key = (inv, round(amt, 2)) if (inv and amt is not None) else None
        if key is not None and key in by_key:
            keep = data_merge_unit(by_key[key], data)
            continue
        data = dict(data)
        if key is not None:
            by_key[key] = data
        out.append(data)
    return out


def data_merge_unit(keep, extra):
    """Append `extra`'s unit onto `keep`'s unit field (semicolon-joined, de-duplicated)."""
    u = (extra.get("unit") or "").strip()
    if not u:
        return keep
    seen = [p.strip() for p in (keep.get("unit") or "").split(";") if p.strip()]
    if u not in seen:
        keep["unit"] = "; ".join(seen + [u]) if seen else u
    return keep


def build_seen_index(wb) -> set:
    """
    Collect the duplicate keys already in the master log, so duplicates are caught across
    runs - not just within a single run. Columns are looked up by header name, so this keeps
    working even if older sheets have a different column layout. The key is rebuilt with the
    same logic write_invoice() uses - including the property/amount disambiguation for
    date-based numbers - so a saved row and a freshly-read invoice compare identically.
    """
    seen = set()
    if MASTER_SHEET not in wb.sheetnames:
        return seen
    ws = wb[MASTER_SHEET]
    header = [c.value for c in ws[1]]
    idx = {name: header.index(name) for name in (
        "Vendor Name", "Invoice #", "Unit", "Invoice Date", "Due Date", "Property", "Amount",
    ) if name in header}
    if "Vendor Name" not in idx or "Invoice #" not in idx:
        return seen

    def cell(row, name):
        i = idx.get(name)
        return row[i] if (i is not None and i < len(row)) else ""

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        invno      = cell(row, "Invoice #")
        date_based = _is_date_based_number(invno, cell(row, "Invoice Date"), cell(row, "Due Date"))
        key = _invoice_key(
            str(cell(row, "Vendor Name") or ""), str(invno or ""), str(cell(row, "Unit") or ""),
            property_name=str(cell(row, "Property") or ""), amount=cell(row, "Amount"),
            date_based=date_based,
        )
        if key:
            seen.add(key)
    return seen


def _parse_amount(value):
    """Parse a money value like '$1,228.50' into a float, or None if not a clean number."""
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    s = re.sub(r"[^0-9.\-]", "", str(value))
    if not s or s in ("-", ".", "-.", "--"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _summarize_line_items(items) -> str:
    """
    Build a short review note from an itemized breakdown. Returns '' unless there are 2+
    line items (a single charge isn't 'itemized'). Caps length so cells stay readable.
    """
    if not isinstance(items, list):
        return ""
    parts = []
    for it in items:
        if isinstance(it, dict):
            desc  = str(it.get("description") or it.get("item") or "").strip()
            amt   = _parse_amount(it.get("amount"))
            amt_s = f"${amt:,.2f}" if amt is not None else str(it.get("amount") or "").strip()
            piece = " - ".join(p for p in (desc, amt_s) if p)
        else:
            piece = str(it).strip()
        if piece:
            parts.append(piece)
    if len(parts) < 2:
        return ""
    note = f"{len(parts)} items: " + " | ".join(parts)
    return note if len(note) <= 500 else note[:497] + "..."


def write_invoice(data: Dict, source_file: str, seen: set, date_processed: str = None,
                  needs_review: bool = False, stored_file: str = ""):
    """
    Record one invoice into the database (the master log). Flags duplicates via `seen` (which is
    updated in place), and sets the needs_review flag when the service location wasn't in the
    property list. Returns (status, vendor, invoice_no) with status 'OK' or 'DUPLICATE'.

    Same identity/dedup and amount logic as the original xlsx writer, so behavior is unchanged -
    only the destination is the DB instead of a worksheet.
    """
    vendor     = (data.get("vendor_name") or "").strip() or "Unknown Vendor"
    invoice_no = resolve_invoice_number(data)
    unit       = (data.get("unit") or "").strip()
    when       = date_processed or datetime.today().strftime("%m/%d/%Y")

    # Amount as a number when possible; otherwise keep the raw text in amount_text.
    parsed_amount = _parse_amount(data.get("total_amount"))
    amount_text   = "" if parsed_amount is not None else (
        str(data.get("total_amount")).strip() if data.get("total_amount") else "")
    items_note    = _summarize_line_items(data.get("line_items"))

    date_based = _is_date_based_number(invoice_no, data.get("invoice_date"), data.get("due_date"))
    key    = _invoice_key(vendor, invoice_no, unit,
                          property_name=(data.get("property") or ""),
                          amount=parsed_amount, date_based=date_based)
    is_dup = bool(key) and key in seen
    status = "DUPLICATE" if is_dup else "OK"
    if key:
        seen.add(key)

    db.insert_invoice({
        "status":         status,
        "vendor_name":    vendor,
        "invoice_number": invoice_no,
        "unit":           unit,
        "invoice_date":   data.get("invoice_date") or "",
        "due_date":       data.get("due_date") or "",
        "amount":         parsed_amount,
        "amount_text":    amount_text,
        "description":    data.get("description") or "",
        "line_items":     items_note,
        "property":       data.get("property") or "",
        "source_file":    source_file,
        "date_processed": when,
        "entered_in_yardi": 0,
        "stored_file":    stored_file,
        "needs_review":   1 if needs_review else 0,
        "origin":         "processor",
    })
    return status, vendor, invoice_no


# Summary-dashboard cell styles
DASH_TITLE_FONT   = Font(bold=True, size=16, name="Arial", color="1F3864")
DASH_SUB_FONT     = Font(italic=True, size=10, name="Arial", color="666666")
DASH_SECTION_FONT = Font(bold=True, size=11, name="Arial", color="FFFFFF")
DASH_LABEL_FONT   = Font(size=11, name="Arial")
DASH_VALUE_FONT   = Font(bold=True, size=12, name="Arial", color="1F3864")
DASH_BANNER_FONT  = Font(bold=True, size=13, name="Arial")
GOOD_FILL = PatternFill("solid", start_color="C6EFCE")            # green - caught up
GOOD_FONT = Font(bold=True, color="006100", name="Arial", size=12)
WARN_FILL = PatternFill("solid", start_color="FFC7CE")            # red - still outstanding
WARN_FONT = Font(bold=True, color="9C0006", name="Arial", size=12)
AMBER_FILL = PatternFill("solid", start_color="FFEB9C")           # amber banner - action needed
AMBER_FONT = Font(bold=True, color="9C6500", name="Arial", size=13)
GOOD_BANNER_FONT = Font(bold=True, color="006100", name="Arial", size=13)


def _init_summary_sheet(ws) -> None:
    """Minimal placeholder. refresh_summary_dashboard() rebuilds this tab into the full
    dashboard at the end of every run; this only avoids a blank tab if a run aborts first."""
    ws["A1"] = "Invoice Tracker — Daily Dashboard"
    ws["A1"].font = DASH_TITLE_FONT
    ws["A2"] = "Run the processor to populate this dashboard."
    ws["A2"].font = DASH_SUB_FONT
    ws.column_dimensions["A"].width = 34


def _countif_literal(text: str) -> str:
    """Escape COUNTIF/SUMIF wildcards so a property name containing * ? ~ matches literally."""
    return re.sub(r"([~*?])", r"~\1", str(text))


def _distinct_properties_in_master(wb) -> list:
    """Property names that have at least one non-duplicate invoice in the master log, in the
    exact spelling stored there (so the dashboard COUNTIFs match), sorted case-insensitively."""
    if MASTER_SHEET not in wb.sheetnames:
        return []
    ws = wb[MASTER_SHEET]
    header = [c.value for c in ws[1]]
    si = header.index("Status")   if "Status"   in header else None
    pi = header.index("Property")  if "Property" in header else None
    if pi is None:
        return []
    names, seen = [], set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        status = row[si] if (si is not None and si < len(row)) else "OK"
        if status == "DUPLICATE":
            continue
        prop = row[pi] if pi < len(row) else None
        name = str(prop).strip() if prop else ""
        if name and name not in seen:
            seen.add(name)
            names.append(name)
    names.sort(key=str.lower)
    return names


def refresh_summary_dashboard(wb, run_date: str) -> None:
    """
    Rebuild the Summary tab into a live "did everything I processed get into Yardi?" board.
    Every count is an Excel FORMULA over the 'All Invoices' log, so it updates the instant you
    tick an 'Entered in Yardi' box in Excel - no need to re-run the script to see progress.
    Only first-time invoices (Status 'OK') count as work to enter; duplicates are ignored.
    """
    _get_or_create_master_sheet(wb)        # guarantee 'All Invoices' exists so formulas resolve

    # Column letters come from MASTER_HEADER so the formulas survive a future column reorder.
    sc = get_column_letter(MASTER_HEADER.index("Status") + 1)
    ac = get_column_letter(MASTER_HEADER.index("Amount") + 1)
    pc = get_column_letter(MASTER_HEADER.index("Property") + 1)
    dc = get_column_letter(MASTER_HEADER.index("Date Processed") + 1)
    yc = get_column_letter(MASTER_HEADER.index("Entered in Yardi") + 1)
    M   = f"'{MASTER_SHEET}'"
    chk = YARDI_CHECK
    col = lambda c: f"{M}!${c}:${c}"                  # whole column (COUNTIF/SUMIFS)
    rng = lambda c: f"{M}!${c}$2:${c}$100000"         # bounded (SUMPRODUCT, for text-date match)

    properties = _distinct_properties_in_master(wb)

    # Rebuild from scratch so a shrinking property list never leaves stale rows behind.
    if SUMMARY_SHEET in wb.sheetnames:
        del wb[SUMMARY_SHEET]
    ws = wb.create_sheet(SUMMARY_SHEET, 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12

    def section(row, text):
        ws.merge_cells(f"A{row}:C{row}")
        c = ws[f"A{row}"]
        c.value, c.font, c.fill = text, DASH_SECTION_FONT, HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 20

    def metric(row, label, formula, *, currency=False):
        a = ws[f"A{row}"]
        a.value, a.font = label, DASH_LABEL_FONT
        a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        b = ws[f"B{row}"]
        b.value, b.font = formula, DASH_VALUE_FONT
        b.alignment = Alignment(horizontal="center", vertical="center")
        if currency:
            b.number_format = '$#,##0.00'
        return b

    ws["A1"] = "Invoice Tracker — Daily Dashboard"
    ws["A1"].font = DASH_TITLE_FONT
    ws.row_dimensions[1].height = 24
    ws["A2"] = f"Last run: {datetime.today().strftime('%B %d, %Y %I:%M %p')}"
    ws["A2"].font = DASH_SUB_FONT

    # Status banner - green when caught up, amber when something is still to enter. It reads
    # the 'still to enter (all dates)' number in B11, so it flips live as you tick boxes.
    ws.merge_cells("A4:C4")
    banner = ws["A4"]
    banner.value = (
        '=IF($B$11=0,'
        '"✓  All caught up — every processed invoice is entered in Yardi",'
        '"⚠  "&$B$11&" invoice(s) still to enter in Yardi (see below)")'
    )
    banner.font = DASH_BANNER_FONT
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 30
    ws.conditional_formatting.add("A4:C4",
        FormulaRule(formula=["$B$11=0"], fill=GOOD_FILL, font=GOOD_BANNER_FONT))
    ws.conditional_formatting.add("A4:C4",
        FormulaRule(formula=["$B$11>0"], fill=AMBER_FILL, font=AMBER_FONT))

    # This run. Date Processed is stored as TEXT, so these use SUMPRODUCT with a text compare;
    # COUNTIF would coerce "06/03/2026" to a date serial and match nothing.
    section(6, f"This run — processed {run_date}")
    metric(7, "Invoices processed",
           f'=SUMPRODUCT(({rng(sc)}="OK")*({rng(dc)}="{run_date}"))')
    today_left = metric(8, "Still to enter in Yardi",
           f'=SUMPRODUCT(({rng(sc)}="OK")*({rng(dc)}="{run_date}")*({rng(yc)}<>"{chk}"))')

    # Overall.
    section(10, "Overall — Yardi entry")
    overall_left = metric(11, "Still to enter (all dates)",
           f'=COUNTIF({col(sc)},"OK")-COUNTIFS({col(sc)},"OK",{col(yc)},"{chk}")')
    metric(12, "Amount still to enter",
           f'=SUMIFS({col(ac)},{col(sc)},"OK")-SUMIFS({col(ac)},{col(sc)},"OK",{col(yc)},"{chk}")',
           currency=True)
    metric(13, "Already entered in Yardi",
           f'=COUNTIFS({col(sc)},"OK",{col(yc)},"{chk}")')
    metric(14, "Total invoices (excl. duplicates)", f'=COUNTIF({col(sc)},"OK")')

    # Color the two "still to enter" numbers: green at 0, red above 0.
    for ref in (today_left.coordinate, overall_left.coordinate):
        ws.conditional_formatting.add(ref,
            CellIsRule(operator="equal", formula=["0"], fill=GOOD_FILL, font=GOOD_FONT))
        ws.conditional_formatting.add(ref,
            CellIsRule(operator="greaterThan", formula=["0"], fill=WARN_FILL, font=WARN_FONT))

    # Per-property breakdown - where the outstanding work is.
    section(16, "Outstanding by property")
    for i, title in enumerate(("Property", "To enter", "Total")):
        c = ws.cell(row=17, column=i + 1, value=title)
        c.font, c.fill = HEADER_FONT, HEADER_FILL
        c.alignment = Alignment(horizontal="left" if i == 0 else "center", vertical="center")
    ws.row_dimensions[17].height = 18

    r = 18
    for name in properties:
        lit = _countif_literal(name)
        a = ws.cell(row=r, column=1, value=name)
        a.font = DASH_LABEL_FONT
        a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        b = ws.cell(row=r, column=2,
            value=f'=COUNTIFS({col(sc)},"OK",{col(pc)},"{lit}")-'
                  f'COUNTIFS({col(sc)},"OK",{col(pc)},"{lit}",{col(yc)},"{chk}")')
        t = ws.cell(row=r, column=3, value=f'=COUNTIFS({col(sc)},"OK",{col(pc)},"{lit}")')
        for cell in (b, t):
            cell.font = DASH_LABEL_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.conditional_formatting.add(b.coordinate,
            CellIsRule(operator="greaterThan", formula=["0"], fill=WARN_FILL, font=WARN_FONT))
        r += 1
    if not properties:
        ws.cell(row=r, column=1, value="(no invoices logged yet)").font = DASH_SUB_FONT


def load_or_create_workbook(path) -> openpyxl.Workbook:
    """Load the existing workbook or create a fresh one. Always ensures a Summary tab."""
    if Path(path).exists():
        wb = openpyxl.load_workbook(path)
        if SUMMARY_SHEET not in wb.sheetnames:
            _init_summary_sheet(wb.create_sheet(SUMMARY_SHEET, 0))
        return wb

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SUMMARY_SHEET
    _init_summary_sheet(ws)
    return wb


def migrate_vendor_tabs_if_needed(wb, path, properties):
    """
    One-time upgrade for spreadsheets from the old layout (one tab per vendor, no master
    log). Re-files those rows into the new property tabs + 'All Invoices' log, backing up
    the original file first. Each row's property is re-validated against `properties`, so an
    old value that isn't a real property lands on the Needs Review tab. Returns
    (migrated_count, backup_path_or_None, review_count).
    """
    if MASTER_SHEET in wb.sheetnames:
        return 0, None, 0                    # already new layout - nothing to do

    old_sheets = [s for s in wb.sheetnames if s not in RESERVED_SHEETS]
    _get_or_create_master_sheet(wb)          # create the master log
    if not old_sheets:
        return 0, None, 0

    # Back up the original file before restructuring it.
    backup = None
    src = Path(path)
    if src.exists():
        stamp  = datetime.today().strftime("%Y%m%d_%H%M%S")
        backup = src.with_name(f"{src.stem}_backup_{stamp}{src.suffix}")
        try:
            shutil.copy2(src, backup)
        except OSError:
            backup = None

    # Old per-vendor columns: Vendor, Invoice#, InvDate, DueDate, Desc, Property, Source, Processed
    seen = set()
    migrated = 0
    reviews  = 0
    for name in old_sheets:
        for row in wb[name].iter_rows(min_row=2, values_only=True):
            if not row or not any(v not in (None, "") for v in row):
                continue
            cells = list(row) + [""] * 8                 # pad short rows
            data = {
                "vendor_name":    cells[0],
                "invoice_number": cells[1],
                "invoice_date":   cells[2],
                "due_date":       cells[3],
                "description":    cells[4],
                "property":       cells[5],
            }
            source_file    = cells[6] or ""
            date_processed = cells[7] or datetime.today().strftime("%m/%d/%Y")

            # Re-validate the stored property against the current list (strict).
            canonical = match_property(data.get("property"), properties)
            tab = None
            if canonical:
                data["property"] = canonical
            elif properties:
                data["property"] = (str(cells[5]).strip() or "(none found)")
                tab = PROPERTY_REVIEW_TAB
                reviews += 1

            write_invoice(wb, data, source_file, seen,
                          date_processed=date_processed, property_tab=tab)
            migrated += 1

    for name in old_sheets:                              # drop the old vendor tabs
        del wb[name]

    return migrated, backup, reviews


# - FILE TYPE DETECTION -

EXTENSION_MAP = {
    ".pdf":  "application/pdf",
    ".jpg":  "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png":  "image/png",
    ".gif":  "image/gif",
    ".webp": "image/webp",
    ".tiff": "image/tiff",
    ".tif":  "image/tiff",
}


def get_media_type_for_path(path: Path) -> Optional[str]:
    """Resolve the Claude media type from a file's extension, or None if unsupported."""
    return EXTENSION_MAP.get(path.suffix.lower())


# - FILE MOVE -

def _safe_folder_name(name: str) -> str:
    """Make a property name safe to use as a Windows folder name."""
    clean = re.sub(r'[<>:"/\\|?*]', "", str(name)).strip().rstrip(". ")
    return clean[:80].strip() or "Unknown Property"


def _build_processed_filename(data: Dict, original: Path, vendor_map=None) -> str:
    """
    Rename a processed invoice to 'VendorName_MM_YYYY' + its original extension. VendorName is
    the vendor's short name from vendors.csv when one matches (e.g. LADWP), otherwise the
    vendor's first word. MM/YYYY come from the invoice date (falling back to the due date).
    """
    vendor = (data.get("vendor_name") or "").strip()
    short  = match_vendor_short_name(vendor, vendor_map)
    if short:
        name = re.sub(r"[^0-9A-Za-z&-]", "", short) or "Vendor"
    else:
        first = re.split(r"\s+", vendor)[0] if vendor else ""
        name  = re.sub(r"[^0-9A-Za-z&-]", "", first) or "Vendor"

    dt = _parse_date(data.get("invoice_date")) or _parse_date(data.get("due_date"))
    base = f"{name}_{dt.strftime('%m')}_{dt.strftime('%Y')}" if dt else name
    return base + original.suffix.lower()


def _unique_name(dest_dir: Path, name: str, reserved: set) -> str:
    """A filename in dest_dir that collides with neither an existing file nor a name already
    reserved this run. Same numeric-suffix rule move_to_processed has always used."""
    target = Path(name)
    stem, suffix = target.stem, target.suffix
    candidate, counter = target.name, 1
    while (dest_dir / candidate).exists() or candidate in reserved:
        candidate = f"{stem}_{counter}{suffix}"
        counter += 1
    reserved.add(candidate)
    return candidate


def reserve_processed_name(subfolder: str, name: str, reserved: set) -> str:
    """
    Claim the final filed name for a file up front - before its rows are written - so the
    name can be recorded in the spreadsheet (the sidecar join key) and the move below lands
    on exactly it. `reserved` carries the names claimed so far this run.
    """
    dest_dir = PROCESSED_FOLDER / _safe_folder_name(subfolder) if subfolder else PROCESSED_FOLDER
    return _unique_name(dest_dir, name, reserved)


def move_to_processed(src: Path, subfolder: str = "", final_name: str = None) -> Path:
    """
    Move a processed file into PROCESSED_FOLDER, under a per-property subfolder when one is
    given, renaming it to `final_name` (already reserved by reserve_processed_name). The
    while-loop is a defensive re-suffix in case something external grabbed the name between
    reservation and the move; normally it never fires. Returns the final destination path.
    """
    dest_dir = PROCESSED_FOLDER / _safe_folder_name(subfolder) if subfolder else PROCESSED_FOLDER
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / (final_name or src.name)
    stem, suffix, counter = dest.stem, dest.suffix, 1
    while dest.exists():
        dest = dest_dir / f"{stem}_{counter}{suffix}"
        counter += 1
    shutil.move(str(src), str(dest))
    return dest


def is_output_locked(path) -> bool:
    """
    True if the spreadsheet exists but can't be opened for writing (typically because it's
    open in Excel). Lets us warn up front instead of failing the save after files have
    already been moved.
    """
    p = Path(path)
    if not p.exists():
        return False
    try:
        with open(p, "r+b"):
            return False
    except (PermissionError, OSError):
        return True


# - AMOUNT SIDECAR (handoff to the Bank Rec Auto-Assembler) -

SIDECAR_NAME   = "_amounts.csv"
SIDECAR_HEADER = ["stored_file", "amount", "vendor", "invoice_number",
                  "unit", "invoice_date", "property", "source_file", "check_number"]


def _write_sidecar(folder: Path, rows: list) -> None:
    with (folder / SIDECAR_NAME).open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=SIDECAR_HEADER)
        w.writeheader()
        w.writerows(rows)


def export_amount_sidecars(wb, processed_root: Path) -> int:
    """
    Refresh processed/<property>/_amounts.csv for the PENDING pool - every non-duplicate,
    not-yet-reconciled invoice that has a filed (stored) name - so the Bank Rec assembler can
    read each invoice's verified amount instead of OCR-ing it. Keyed on the stored filename
    the assembler sees; a multi-bill PDF appears as several rows sharing one stored_file. A
    property whose pool is now empty gets a header-only sidecar so already-reconciled files
    stop being re-staged. Returns the number of sidecars written with pending rows.
    """
    if MASTER_SHEET not in wb.sheetnames:
        return 0
    ws = wb[MASTER_SHEET]
    header = [c.value for c in ws[1]]
    idx = {n: header.index(n) for n in (
        "Status", "Stored File", "Amount", "Vendor Name", "Invoice #", "Unit",
        "Invoice Date", "Property", "Source File", "Reconciled", "Check #") if n in header}
    if "Stored File" not in idx or "Property" not in idx:
        return 0   # nothing to key on yet (e.g. an old sheet before ensure_columns ran)

    def cell(row, name):
        i = idx.get(name)
        return row[i] if (i is not None and i < len(row)) else ""

    by_safe = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or str(cell(row, "Status") or "OK") == "DUPLICATE":
            continue
        if str(cell(row, "Reconciled") or "").strip():        # already reconciled -> not pending
            continue
        stored = str(cell(row, "Stored File") or "").strip()
        prop   = str(cell(row, "Property") or "").strip()
        if not stored or not prop:
            continue
        amt = _parse_amount(cell(row, "Amount"))
        by_safe.setdefault(_safe_folder_name(prop), []).append({
            "stored_file":    stored,
            "amount":         f"{amt:.2f}" if amt is not None else "",
            "vendor":         str(cell(row, "Vendor Name") or "").strip(),
            "invoice_number": str(cell(row, "Invoice #") or "").strip(),
            "unit":           str(cell(row, "Unit") or "").strip(),
            "invoice_date":   str(cell(row, "Invoice Date") or "").strip(),
            "property":       prop,
            "source_file":    str(cell(row, "Source File") or "").strip(),
            "check_number":   re.sub(r"[^0-9]", "", str(cell(row, "Check #") or "")),
        })

    written, refreshed = 0, set()
    for safe, rows in by_safe.items():
        folder = processed_root / safe
        if not folder.exists():
            continue
        try:
            _write_sidecar(folder, rows)
            written += 1
            refreshed.add(folder.resolve())
        except OSError as exc:
            print(f"  [!] Could not write {SIDECAR_NAME} for '{safe}' ({exc})")
    for sc in processed_root.glob("*/" + SIDECAR_NAME):        # empty out stale sidecars
        if sc.parent.resolve() not in refreshed:
            try:
                _write_sidecar(sc.parent, [])
            except OSError:
                pass
    return written


# - PROPERTY NORMALIZATION -

# How close a fuzzy match must be (0.0-1.0) before we trust it enough to normalize.
# High on purpose: only near-identical spellings auto-map; anything iffy is left as-is.
PROPERTY_FUZZY_THRESHOLD = 0.86

# Tab where invoices whose service location isn't in properties.csv are parked for review,
# instead of recording a billing/management address that isn't a real property.
PROPERTY_REVIEW_TAB = "Needs Review"


def _normalize_key(text: str) -> str:
    """Lowercase and drop everything but letters/digits, so 'J.H. Lee' == 'JH Lee'."""
    return re.sub(r"[^a-z0-9]", "", text.lower())


def load_property_list(path=None) -> list:
    """
    Property list as [(canonical_name, {normalized_keys}, [alias_text, ...]), ...] - now sourced
    from the database (the DB is the single source of truth). `path` is ignored, kept so existing
    call sites (which pass PROPERTY_LIST_FILE) work unchanged.
    """
    return db.load_property_list()


def build_property_reference(properties) -> str:
    """Render the property list for the extraction prompt: one line per property with its
    addresses/aliases, so Claude can match an invoice's service location to a canonical name."""
    lines = []
    for canonical, _keys, aliases in properties:
        if aliases:
            lines.append(f"- {canonical}  (addresses/aliases: {'; '.join(aliases)})")
        else:
            lines.append(f"- {canonical}")
    return "\n".join(lines)


def match_property(raw, properties) -> Optional[str]:
    """
    Match an extracted service location to a canonical property name from the list.
    Returns the canonical name, or None when there's no confident match (or no list) - so
    callers can enforce that a property comes strictly from properties.csv.
    """
    if raw is None or not properties:
        return None
    raw_str = str(raw).strip()
    if not raw_str:
        return None
    raw_key = _normalize_key(raw_str)
    if not raw_key:
        return None

    # 1) Exact match on a canonical name or alias (ignoring case/spacing/punctuation).
    for canonical, keys, _aliases in properties:
        if raw_key in keys:
            return canonical

    # 2) A known alias appears *inside* the extracted text - common when the invoice
    #    prints a full street address. Prefer the longest alias as the most specific.
    if len(raw_key) >= 5:
        best_canonical, best_len = None, 0
        for canonical, keys, _aliases in properties:
            for key in keys:
                if len(key) >= 5 and key in raw_key and len(key) > best_len:
                    best_canonical, best_len = canonical, len(key)
        if best_canonical:
            return best_canonical

    # 3) Fuzzy fallback for near-miss spellings / typos.
    best_canonical, best_score = None, 0.0
    for canonical, keys, _aliases in properties:
        for key in keys:
            score = difflib.SequenceMatcher(None, raw_key, key).ratio()
            if score > best_score:
                best_canonical, best_score = canonical, score
    if best_canonical and best_score >= PROPERTY_FUZZY_THRESHOLD:
        return best_canonical

    # No confident match.
    return None


def resolve_property(query, properties) -> list:
    """
    Forgiving property lookup for command-line use (the strict match_property is for reading
    invoices). Matches `query` against each property's canonical name, code, and aliases:
    exact (normalized) first, then a substring match either direction - so "1711 N" finds the
    property whose alias is "1711 N Alexandria Ave", "beach" finds "6281-6301 Beach Blvd", a
    short code finds its property, etc. Returns the list of matching canonical names (0, 1, or
    several when the query is ambiguous).
    """
    q = _normalize_key(query or "")
    if not q:
        return []
    exact = [c for c, keys, _a in properties if q in keys]
    if exact:
        return list(dict.fromkeys(exact))
    hits = [c for c, keys, _a in properties if any(q in k or k in q for k in keys)]
    return list(dict.fromkeys(hits))


def load_vendor_map(path=None) -> list:
    """
    Vendor short-name map as [(short_name, {normalized_keys}), ...], sourced from the database.
    `path` is ignored (kept for call-site compatibility).
    """
    return db.load_vendor_map()


def match_vendor_short_name(vendor, vendor_map) -> Optional[str]:
    """
    Return a short name for the vendor from vendors.csv (for filenames only), or None.
    Matches on exact name, an alias appearing inside the vendor name, then a close spelling.
    """
    if not vendor or not vendor_map:
        return None
    vkey = _normalize_key(str(vendor))
    if not vkey:
        return None
    for short, keys in vendor_map:                     # 1) exact
        if vkey in keys:
            return short
    best, best_len = None, 0                           # 2) alias inside the vendor name
    for short, keys in vendor_map:
        for key in keys:
            if len(key) >= 4 and key in vkey and len(key) > best_len:
                best, best_len = short, len(key)
    if best:
        return best
    best, best_score = None, 0.0                       # 3) fuzzy
    for short, keys in vendor_map:
        for key in keys:
            score = difflib.SequenceMatcher(None, vkey, key).ratio()
            if score > best_score:
                best, best_score = short, score
    return best if best and best_score >= PROPERTY_FUZZY_THRESHOLD else None


# - MANUAL ENTRY (fallback when auto-extraction fails) -

def prompt_manual_invoice(path, properties):
    """
    Fallback for a file the auto-extractor couldn't read: let the user type the invoice in by
    hand. Returns (data_dict, filed_name_override_or_None) shaped exactly like the extractor's
    output - so it flows through the same logging / filing / sidecar pipeline - or None to skip.
    Also returns None when there's no interactive console, so unattended runs never hang.
    """
    try:
        if not (sys.stdin and sys.stdin.isatty()):
            return None
    except Exception:
        return None

    print(f"      Auto-extraction failed for '{path.name}'.")
    if input("      Enter this invoice by hand instead? [y/N]: ").strip().lower() not in ("y", "yes"):
        return None

    def ask(label, required=False):
        while True:
            v = input(f"        {label}: ").strip()
            if v or not required:
                return v
            print("        (this one is required)")

    vendor = ask("Vendor name", required=True)
    amount = ask("Total amount (e.g. 1387.74)", required=True)

    if properties:
        print("        Known properties:")
        for i, (canon, _keys, _aliases) in enumerate(properties, 1):
            print(f"          {i:>2}. {canon}")
        praw = ask("Property (type a name or its number)", required=True)
        prop = (properties[int(praw) - 1][0]
                if (praw.isdigit() and 1 <= int(praw) <= len(properties)) else praw)
    else:
        prop = ask("Property", required=True)

    invno = ask("Invoice # (blank to use the date)")
    unit  = ask("Unit (blank if none)")
    idate = ask("Invoice date MM/DD/YYYY (blank if unknown)")
    desc  = ask("Short description (optional)")
    fname = ask("Filed name (blank = auto, e.g. SoCalGas_06_2026.pdf)")

    data = {
        "vendor_name":    vendor,
        "total_amount":   amount,
        "property":       prop,
        "invoice_number": invno or None,
        "unit":           unit or None,
        "invoice_date":   idate or None,
        "due_date":       None,
        "description":    desc or None,
        "line_items":     [],
    }
    return data, (fname or None)


# - MAIN -

def main():
    sep = "=" * 55
    print(sep)
    print("  Invoice Processor  |  Folder -> Claude -> Database")
    print(sep)

    anthropic_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not anthropic_key or "xxxx" in anthropic_key.lower():
        print("\n[ERROR] Your ANTHROPIC_API_KEY isn't filled in yet.")
        print("  Add it to the .env file as  ANTHROPIC_API_KEY=sk-ant-...  (get a key at")
        print("  https://console.anthropic.com), then run again.\n")
        _pause()
        sys.exit(1)

    db.init()   # ensure the database and tables exist

    # First run: create the input folder so the user has somewhere to drop files.
    if not INPUT_FOLDER.exists():
        INPUT_FOLDER.mkdir(parents=True, exist_ok=True)
        print(f"\n[*] Created the input folder:\n      {INPUT_FOLDER}")
        print("    Put your invoice PDFs/images in there, then run this again.\n")
        _pause()
        return

    # Gather files from the top level of the input folder, splitting supported
    # invoices from anything we can't read (so the user isn't left guessing).
    entries       = [p for p in INPUT_FOLDER.iterdir() if p.is_file()]
    files         = sorted(p for p in entries if get_media_type_for_path(p))
    skipped_types = sorted(p.name for p in entries if not get_media_type_for_path(p))

    if not files:
        print(f"\n[*] No invoices found in:\n      {INPUT_FOLDER}")
        print("    Drop your PDF or image invoices in there and run again.")
        if skipped_types:
            print(f"    (Ignoring unsupported file(s): {', '.join(skipped_types)})")
        print()
        _pause()
        return

    print(f"\n[*] Found {len(files)} invoice file(s) in '{INPUT_FOLDER.name}'")

    claude = anthropic.Anthropic(api_key=anthropic_key)

    properties = load_property_list()
    if properties:
        word = "property" if len(properties) == 1 else "properties"
        print(f"[*] Loaded {len(properties)} {word} from the database; "
              f"service locations will be matched to this list")
    else:
        print("[*] No properties in the database - recording the service location text as-is")
    extraction_prompt = build_extraction_prompt(build_property_reference(properties))

    # Optional vendor short-name map (used only to shorten the processed filename).
    vendor_map = load_vendor_map()
    if vendor_map:
        print(f"[*] Loaded {len(vendor_map)} vendor short-name(s) from the database")

    # Index of invoices already recorded, so duplicates are caught across runs.
    seen = db.seen_index()

    total_logged = 0
    duplicates   = 0
    moved        = 0
    normalized   = 0
    unmatched    = 0
    failed_files = []
    to_move      = []
    reserved     = set()    # filed names claimed this run (prevents in-run name collisions)

    for path in files:
        print(f"\n[FILE] {path.name}")
        media_type = get_media_type_for_path(path)

        try:
            file_bytes = path.read_bytes()
        except OSError as exc:
            print(f"  [!] Could not read file: {exc}")
            failed_files.append(path.name)
            continue

        if not file_bytes:
            print("  [!] Empty file - leaving it in place")
            failed_files.append(path.name)
            continue

        try:
            invoices = extract_invoice_data(claude, file_bytes, media_type, extraction_prompt)
        except anthropic.AuthenticationError:
            print("\n[ERROR] Claude rejected your API key (authentication failed).")
            print("  Fix ANTHROPIC_API_KEY in your .env (paste a fresh key from")
            print("  https://console.anthropic.com), then run again.")
            print(f"  Nothing was changed - every file is still in '{INPUT_FOLDER.name}'.\n")
            _pause()
            sys.exit(1)
        manual_name = None
        if not invoices:
            manual = prompt_manual_invoice(path, properties)
            if manual is None:
                print("  [!] Could not extract invoice data - leaving file in place "
                      "(run again and choose manual entry, or fix the file)")
                failed_files.append(path.name)
                continue
            invoices, manual_name = [manual[0]], manual[1]
            print("  Recording your manual entry")

        # One invoice the extractor split per line-item/unit (same invoice # + same total on
        # each element) is folded back into a single row so its amount isn't counted per unit.
        before = len(invoices)
        invoices = merge_oversplit_invoices(invoices)
        if len(invoices) < before:
            print(f"  Merged {before - len(invoices)} duplicated line(s) of one invoice "
                  f"(same invoice # and total) into a single row")

        # A file may hold several bills (e.g. one per unit of a property) - record each.
        if len(invoices) > 1:
            print(f"  Found {len(invoices)} invoices in this file")

        file_folder = None     # the whole file is filed under the first invoice's property
        file_name   = None
        stored_name = ""       # the reserved filed name, recorded on every row of this file
        for data in invoices:
            # Property must come strictly from properties.csv: it's the service location,
            # matched to a canonical name. Anything unmatched is parked on "Needs Review".
            raw_property = data.get("property")
            canonical    = match_property(raw_property, properties)
            review       = False
            if canonical:
                if _normalize_key(str(raw_property or "")) != _normalize_key(canonical):
                    print(f"      Property matched: '{raw_property}' -> '{canonical}'")
                data["property"] = canonical
                normalized += 1
            elif properties:
                # Strict mode: keep what Claude read for reference, but flag it for review.
                data["property"] = str(raw_property).strip() if raw_property else "(none found)"
                review = True
                unmatched += 1
                print(f"  [REVIEW] Service location not in properties.csv "
                      f"(read: '{raw_property or '—'}') -> '{PROPERTY_REVIEW_TAB}' tab")
            else:
                # No property list loaded - record whatever was read.
                data["property"] = str(raw_property).strip() if raw_property else ""

            # Decide the file's folder + final stored name from the FIRST invoice, before its
            # row is written, so every row of this file carries the same join key the
            # assembler will see. The name is reserved now; the actual move happens after the
            # save (below), to exactly this name.
            if file_folder is None:
                file_folder = PROPERTY_REVIEW_TAB if review else (data.get("property") or "Unknown Property")
                if manual_name:                      # user-supplied filed name (keep ext sane)
                    intended = manual_name if os.path.splitext(manual_name)[1] else manual_name + path.suffix.lower()
                else:
                    intended = _build_processed_filename(data, path, vendor_map)
                stored_name = reserve_processed_name(file_folder, intended, reserved)
                file_name   = stored_name

            status, vendor, inv_no = write_invoice(
                data, path.name, seen,
                needs_review=review,
                stored_file=stored_name,
            )
            total_logged += 1
            unit   = (data.get("unit") or "").strip()
            unit_s = f"  Unit: {unit}" if unit else ""
            dest_prop = PROPERTY_REVIEW_TAB if review else (data.get("property") or "Unknown Property")
            if status == "DUPLICATE":
                duplicates += 1
                print(f"  [DUPLICATE]  {vendor}  |  Invoice #: {inv_no or '--'}{unit_s} "
                      f"- already recorded")
            else:
                print(f"  OK  {vendor}  |  Invoice #: {inv_no or '--'}{unit_s}  ->  {dest_prop}")

        # Defer the move until the spreadsheet is safely saved (below), so a failed save
        # never strands a file in 'processed' without its rows recorded.
        to_move.append((path, file_folder, file_name))

    # Each invoice was committed to the DB as it was recorded (write_invoice), so the data is
    # already safe. Now move each processed file into its property's subfolder, renamed to
    # VendorName_MM_YYYY.
    for path, dest_folder, new_name in to_move:
        try:
            dest = move_to_processed(path, dest_folder, new_name)
            moved += 1
            print(f"  Filed: {path.name}  ->  {dest_folder}\\{dest.name}")
        except OSError as exc:
            print(f"  [!] Recorded '{path.name}' but couldn't move it ({exc}).")
            print(f"      Move it out of '{INPUT_FOLDER.name}' yourself to avoid a duplicate next run.")

    # Refresh the per-property amount sidecars (the verified-amount handoff to the Bank Rec
    # assembler). Sourced from the DB; the files are now in processed/<property>/.
    sidecars = db.export_amount_sidecars(PROCESSED_FOLDER)

    print("\n" + sep)
    print(f"  Done!  {total_logged} invoice(s) recorded to the database")
    print(f"         {moved} file(s) moved to '{PROCESSED_FOLDER.name}\\<property>'")
    if sidecars:
        print(f"         {sidecars} amount sidecar(s) refreshed ('{SIDECAR_NAME}' per property)")
    if duplicates:
        print(f"         {duplicates} duplicate(s) flagged")
    if normalized:
        print(f"         {normalized} propert{'y' if normalized == 1 else 'ies'} matched to your list")
    if unmatched:
        print(f"         {unmatched} invoice(s) need a property review -> Needs Review")
    if failed_files:
        print(f"  Note:  {len(failed_files)} file(s) could not be processed and were left")
        print(f"         in '{INPUT_FOLDER.name}': {', '.join(failed_files)}")
    if skipped_types:
        print(f"         Ignored non-invoice file(s): {', '.join(skipped_types)}")
    print(sep)
    _pause("\nPress Enter to close...")   # Keeps window open if double-clicked from a terminal


if __name__ == "__main__":
    main()
