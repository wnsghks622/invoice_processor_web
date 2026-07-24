# -*- coding: utf-8 -*-
"""
core/db.py - the SQLite data layer. Single source of truth for invoices, properties, and
vendors (replaces invoices.xlsx + properties.csv + vendors.csv).

Design notes:
- One short-lived connection per operation, so it's safe under Flask's threaded dev server.
- Pure, storage-independent helpers (dedup key, amount parsing, folder-name safety) are reused
  from processor.py so a row saved here dedups and files identically to the old xlsx path.
- `export_amount_sidecars()` writes byte-identical `_amounts.csv` files, so bankrec.py and
  stage_month.py keep working unchanged.

processor.py imports this module (Phase B). To keep that import cycle-safe, nothing here
touches `ip.*` at module-load time - the processor helpers are only used inside functions.
"""
from __future__ import annotations

import sqlite3
from pathlib import Path
from typing import Optional

import config
from . import processor as ip   # pure helpers only; never called at import time (see note above)

DB_PATH = config.DB_PATH

# Full invoice column set, in order. One place to add a column.
INVOICE_COLUMNS = [
    "status",           # 'OK' | 'DUPLICATE'
    "vendor_name",
    "invoice_number",
    "unit",
    "invoice_date",
    "due_date",
    "amount",           # REAL, parsed; NULL when unparseable (then amount_text holds raw)
    "amount_text",      # original text when amount couldn't be parsed to a number
    "description",
    "line_items",       # the summarized "N items: ..." note (was 'Line Items (review)')
    "property",         # canonical name, or the raw read-text when needs_review=1
    "source_file",      # original dropped filename
    "date_processed",
    "entered_in_yardi", # 0 | 1
    "stored_file",      # filed VendorShort_MM_YYYY.pdf - the sidecar/assembler join key
    "reconciled",       # "<Month YYYY>" once it clears a statement, else ''
    "check_number",
    "carried_forward",  # month it was reviewed but didn't clear, else ''
    "needs_review",     # 0 | 1 (service location not in the property list)
    "origin",           # provenance: 'processor' | 'master' | 'tab:<name>' | 'manual'
]

_SCHEMA = """
CREATE TABLE IF NOT EXISTS properties (
    id             INTEGER PRIMARY KEY AUTOINCREMENT,
    canonical_name TEXT NOT NULL UNIQUE,
    property_code  TEXT DEFAULT '',
    aliases        TEXT DEFAULT '',        -- semicolon-separated, original text
    sort_order     INTEGER DEFAULT 0
);
CREATE TABLE IF NOT EXISTS vendors (
    id         INTEGER PRIMARY KEY AUTOINCREMENT,
    short_name TEXT NOT NULL UNIQUE,
    aliases    TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS invoices (
    id               INTEGER PRIMARY KEY AUTOINCREMENT,
    status           TEXT DEFAULT 'OK',
    vendor_name      TEXT DEFAULT '',
    invoice_number   TEXT DEFAULT '',
    unit             TEXT DEFAULT '',
    invoice_date     TEXT DEFAULT '',
    due_date         TEXT DEFAULT '',
    amount           REAL,
    amount_text      TEXT DEFAULT '',
    description      TEXT DEFAULT '',
    line_items       TEXT DEFAULT '',
    property         TEXT DEFAULT '',
    source_file      TEXT DEFAULT '',
    date_processed   TEXT DEFAULT '',
    entered_in_yardi INTEGER DEFAULT 0,
    stored_file      TEXT DEFAULT '',
    reconciled       TEXT DEFAULT '',
    check_number     TEXT DEFAULT '',
    carried_forward  TEXT DEFAULT '',
    needs_review     INTEGER DEFAULT 0,
    origin           TEXT DEFAULT 'processor'
);
CREATE INDEX IF NOT EXISTS idx_invoices_property   ON invoices(property);
CREATE INDEX IF NOT EXISTS idx_invoices_stored     ON invoices(stored_file);
CREATE INDEX IF NOT EXISTS idx_invoices_reconciled ON invoices(reconciled);
"""


# --------------------------------------------------------------------------- connection

def _connect() -> sqlite3.Connection:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init() -> None:
    """Create tables/indexes if absent. Idempotent."""
    with _connect() as conn:
        conn.executescript(_SCHEMA)


def is_empty() -> bool:
    """True if there are no invoices yet (used by the migration's idempotency guard)."""
    if not DB_PATH.exists():
        return True
    with _connect() as conn:
        return conn.execute("SELECT COUNT(*) FROM invoices").fetchone()[0] == 0


# --------------------------------------------------------------------------- properties

def add_property(canonical_name: str, property_code: str = "", aliases: str = "",
                 sort_order: int = 0) -> int:
    with _connect() as conn:
        cur = conn.execute(
            "INSERT INTO properties (canonical_name, property_code, aliases, sort_order) "
            "VALUES (?,?,?,?)",
            (canonical_name.strip(), property_code.strip(), aliases.strip(), sort_order),
        )
        return cur.lastrowid


def update_property(prop_id: int, canonical_name: str, property_code: str, aliases: str) -> None:
    with _connect() as conn:
        conn.execute(
            "UPDATE properties SET canonical_name=?, property_code=?, aliases=? WHERE id=?",
            (canonical_name.strip(), property_code.strip(), aliases.strip(), prop_id),
        )


def delete_property(prop_id: int) -> None:
    with _connect() as conn:
        conn.execute("DELETE FROM properties WHERE id=?", (prop_id,))


def all_properties() -> list[dict]:
    """Display shape: [{id, name, code, aliases}], in sort order then name."""
    with _connect() as conn:
        rows = conn.execute(
            "SELECT id, canonical_name, property_code, aliases FROM properties "
            "ORDER BY sort_order, canonical_name COLLATE NOCASE"
        ).fetchall()
    return [{"id": r["id"], "name": r["canonical_name"], "code": r["property_code"] or "",
             "aliases": r["aliases"] or ""} for r in rows]


def load_property_list(path=None) -> list:
    """
    DB-backed drop-in for processor.load_property_list(). Returns
    [(canonical_name, {normalized_keys}, [alias_text, ...]), ...] - the exact shape
    match_property()/resolve_property()/build_property_reference() consume. `path` is ignored
    (kept so existing call sites pass through unchanged).
    """
    entries = []
    for p in all_properties():
        canonical, code, aliases_text = p["name"], p["code"], p["aliases"]
        aliases = [a.strip() for a in aliases_text.split(";") if a.strip()]
        keys = {ip._normalize_key(n) for n in [canonical, code, *aliases] if ip._normalize_key(n)}
        if keys:
            entries.append((canonical, keys, aliases))
    return entries


# --------------------------------------------------------------------------- vendors

def add_vendor(short_name: str, aliases: str = "") -> int:
    with _connect() as conn:
        cur = conn.execute("INSERT INTO vendors (short_name, aliases) VALUES (?,?)",
                           (short_name.strip(), aliases.strip()))
        return cur.lastrowid


def update_vendor(vendor_id: int, short_name: str, aliases: str) -> None:
    with _connect() as conn:
        conn.execute("UPDATE vendors SET short_name=?, aliases=? WHERE id=?",
                     (short_name.strip(), aliases.strip(), vendor_id))


def delete_vendor(vendor_id: int) -> None:
    with _connect() as conn:
        conn.execute("DELETE FROM vendors WHERE id=?", (vendor_id,))


def all_vendors() -> list[dict]:
    with _connect() as conn:
        rows = conn.execute(
            "SELECT id, short_name, aliases FROM vendors ORDER BY short_name COLLATE NOCASE"
        ).fetchall()
    return [{"id": r["id"], "short_name": r["short_name"], "aliases": r["aliases"] or ""}
            for r in rows]


def load_vendor_map(path=None) -> list:
    """DB-backed drop-in for processor.load_vendor_map(): [(short_name, {normalized_keys}), ...]."""
    entries = []
    for v in all_vendors():
        short, aliases_text = v["short_name"], v["aliases"]
        aliases = [a.strip() for a in aliases_text.split(";") if a.strip()]
        keys = {ip._normalize_key(n) for n in [short, *aliases] if ip._normalize_key(n)}
        if keys:
            entries.append((short, keys))
    return entries


# --------------------------------------------------------------------------- invoices

def insert_invoice(rec: dict) -> int:
    """Insert one invoice. `rec` may set any of INVOICE_COLUMNS; the rest take table defaults."""
    cols = [c for c in INVOICE_COLUMNS if c in rec]
    placeholders = ",".join("?" for _ in cols)
    values = [rec[c] for c in cols]
    with _connect() as conn:
        cur = conn.execute(
            f"INSERT INTO invoices ({','.join(cols)}) VALUES ({placeholders})", values
        )
        return cur.lastrowid


def seen_index() -> set:
    """
    Duplicate keys already recorded, rebuilt with the SAME logic write_invoice() uses, so a
    freshly-read invoice compares identically to a stored row (mirrors build_seen_index()).
    """
    seen = set()
    with _connect() as conn:
        rows = conn.execute(
            "SELECT vendor_name, invoice_number, unit, invoice_date, due_date, property, "
            "amount, amount_text FROM invoices"
        ).fetchall()
    for r in rows:
        invno = r["invoice_number"]
        date_based = ip._is_date_based_number(invno, r["invoice_date"], r["due_date"])
        amount = r["amount"] if r["amount"] is not None else r["amount_text"]
        key = ip._invoice_key(
            str(r["vendor_name"] or ""), str(invno or ""), str(r["unit"] or ""),
            property_name=str(r["property"] or ""), amount=amount, date_based=date_based,
        )
        if key:
            seen.add(key)
    return seen


def list_invoices(property: Optional[str] = None, needs_review: Optional[bool] = None,
                  pending_only: bool = False, duplicates_only: bool = False,
                  search: str = "") -> list[dict]:
    """Invoices as dicts, newest id first, with optional filters (for the web UI)."""
    where, params = [], []
    if property is not None:
        where.append("property = ?"); params.append(property)
    if needs_review is not None:
        where.append("needs_review = ?"); params.append(1 if needs_review else 0)
    if duplicates_only:
        where.append("status = 'DUPLICATE'")
    if pending_only:
        where.append("status != 'DUPLICATE' AND COALESCE(reconciled,'') = '' "
                     "AND COALESCE(stored_file,'') != ''")
    if search:
        where.append("(vendor_name LIKE ? OR invoice_number LIKE ? OR property LIKE ?)")
        params += [f"%{search}%"] * 3
    sql = "SELECT * FROM invoices"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY id DESC"
    with _connect() as conn:
        return [dict(r) for r in conn.execute(sql, params).fetchall()]


def get_invoice(invoice_id: int) -> Optional[dict]:
    with _connect() as conn:
        r = conn.execute("SELECT * FROM invoices WHERE id=?", (invoice_id,)).fetchone()
    return dict(r) if r else None


def update_invoice(invoice_id: int, **fields) -> None:
    """Update arbitrary columns of one invoice (web edit). Only known columns are applied."""
    cols = [c for c in fields if c in INVOICE_COLUMNS]
    if not cols:
        return
    assignments = ",".join(f"{c}=?" for c in cols)
    values = [fields[c] for c in cols] + [invoice_id]
    with _connect() as conn:
        conn.execute(f"UPDATE invoices SET {assignments} WHERE id=?", values)


def delete_invoice(invoice_id: int) -> None:
    """Remove one invoice record. Does NOT touch the filed PDF on disk - deleting the record is
    reversible by re-processing the file, deleting the file wouldn't be."""
    with _connect() as conn:
        conn.execute("DELETE FROM invoices WHERE id=?", (invoice_id,))


def set_entered_in_yardi(invoice_id: int, entered: bool) -> None:
    with _connect() as conn:
        conn.execute("UPDATE invoices SET entered_in_yardi=? WHERE id=?",
                     (1 if entered else 0, invoice_id))


def counts() -> dict:
    """Headline counts for the dashboard."""
    with _connect() as conn:
        c = conn.cursor()
        total = c.execute("SELECT COUNT(*) FROM invoices").fetchone()[0]
        ok = c.execute("SELECT COUNT(*) FROM invoices WHERE status!='DUPLICATE'").fetchone()[0]
        dup = c.execute("SELECT COUNT(*) FROM invoices WHERE status='DUPLICATE'").fetchone()[0]
        recon = c.execute("SELECT COUNT(*) FROM invoices WHERE COALESCE(reconciled,'')!=''").fetchone()[0]
        review = c.execute("SELECT COUNT(*) FROM invoices WHERE needs_review=1").fetchone()[0]
        yardi = c.execute("SELECT COUNT(*) FROM invoices WHERE entered_in_yardi=1 "
                          "AND status!='DUPLICATE'").fetchone()[0]
        pending = c.execute(
            "SELECT COUNT(*) FROM invoices WHERE status!='DUPLICATE' "
            "AND COALESCE(reconciled,'')='' AND COALESCE(stored_file,'')!=''").fetchone()[0]
    return {"total": total, "ok": ok, "duplicate": dup, "reconciled": recon,
            "needs_review": review, "entered_in_yardi": yardi, "pending": pending}


def pending_by_property() -> dict[str, int]:
    """{canonical_property: pending_invoice_count} (non-dup, not reconciled, has a stored file)."""
    with _connect() as conn:
        rows = conn.execute(
            "SELECT property, COUNT(*) AS n FROM invoices "
            "WHERE status!='DUPLICATE' AND COALESCE(reconciled,'')='' "
            "AND COALESCE(stored_file,'')!='' GROUP BY property"
        ).fetchall()
    return {r["property"]: r["n"] for r in rows}


# ------------------------------------------------------ reconcile / reassign / cleanup support

def mark_reconciled(property: str, stored_file: str, period: str) -> int:
    """Stamp reconciled=<period> and clear carried_forward for matching non-dup rows that
    aren't already reconciled. Matched case-insensitively on (property, stored_file). Returns
    rows updated."""
    with _connect() as conn:
        cur = conn.execute(
            "UPDATE invoices SET reconciled=?, carried_forward='' "
            "WHERE status!='DUPLICATE' AND COALESCE(reconciled,'')='' "
            "AND lower(property)=lower(?) AND lower(stored_file)=lower(?)",
            (period, property, stored_file),
        )
        return cur.rowcount


def set_carried_forward(property: str, stored_file: str, period: str) -> int:
    """Flag rows that were in a rec but didn't clear (only when not already reconciled)."""
    with _connect() as conn:
        cur = conn.execute(
            "UPDATE invoices SET carried_forward=? "
            "WHERE status!='DUPLICATE' AND COALESCE(reconciled,'')='' "
            "AND lower(property)=lower(?) AND lower(stored_file)=lower(?)",
            (period, property, stored_file),
        )
        return cur.rowcount


def reassign_property(invoice_id: int, new_property: str, new_stored_file: str = None) -> None:
    """Move a Needs-Review invoice to its real property and clear the flag."""
    with _connect() as conn:
        if new_stored_file is not None:
            conn.execute("UPDATE invoices SET property=?, needs_review=0, stored_file=? WHERE id=?",
                         (new_property, new_stored_file, invoice_id))
        else:
            conn.execute("UPDATE invoices SET property=?, needs_review=0 WHERE id=?",
                         (new_property, invoice_id))


def review_invoices() -> list[dict]:
    """All invoices still on Needs Review (for reassign_review)."""
    return list_invoices(needs_review=True)


def file_status() -> tuple[set, set]:
    """
    For cleanup: (reconciled, dup_only), each a set of (normalized_property, stored_file_lower).
      reconciled = the file has a reconciled mark (cleared)
      dup_only   = every row for the file is a DUPLICATE (no OK row, not reconciled)
    Mirrors cleanup_processed.file_status().
    """
    info = {}
    with _connect() as conn:
        rows = conn.execute(
            "SELECT status, reconciled, property, stored_file FROM invoices"
        ).fetchall()
    for r in rows:
        sf = str(r["stored_file"] or "").strip()
        if not sf:
            continue
        k = (ip._normalize_key(str(r["property"] or "")), sf.lower())
        d = info.setdefault(k, {"ok": False, "dup": False, "rec": False})
        d["rec"] = d["rec"] or bool(str(r["reconciled"] or "").strip())
        if str(r["status"] or "OK").strip() == "DUPLICATE":
            d["dup"] = True
        else:
            d["ok"] = True
    reconciled = {k for k, d in info.items() if d["rec"]}
    dup_only = {k for k, d in info.items() if d["dup"] and not d["ok"] and not d["rec"]}
    return reconciled, dup_only


# --------------------------------------------------------------------------- amount sidecars

def export_amount_sidecars(processed_root: Path = None) -> int:
    """
    Refresh processed/<property>/_amounts.csv for the PENDING pool - byte-compatible with the
    old xlsx exporter (same SIDECAR_HEADER), so bankrec.py / stage_month.py are unaffected.
    A property whose pool is now empty gets a header-only sidecar so cleared files stop being
    re-staged. Returns the number of sidecars written with pending rows.
    """
    processed_root = Path(processed_root or config.PROCESSED)
    with _connect() as conn:
        rows = conn.execute(
            "SELECT stored_file, amount, amount_text, vendor_name, invoice_number, unit, "
            "invoice_date, property, source_file, check_number FROM invoices "
            "WHERE status!='DUPLICATE' AND COALESCE(reconciled,'')='' "
            "AND COALESCE(stored_file,'')!='' AND COALESCE(property,'')!=''"
        ).fetchall()

    import re
    by_safe: dict[str, list] = {}
    for r in rows:
        amt = ip._parse_amount(r["amount"] if r["amount"] is not None else r["amount_text"])
        by_safe.setdefault(ip._safe_folder_name(str(r["property"])), []).append({
            "stored_file":    str(r["stored_file"] or "").strip(),
            "amount":         f"{amt:.2f}" if amt is not None else "",
            "vendor":         str(r["vendor_name"] or "").strip(),
            "invoice_number": str(r["invoice_number"] or "").strip(),
            "unit":           str(r["unit"] or "").strip(),
            "invoice_date":   str(r["invoice_date"] or "").strip(),
            "property":       str(r["property"] or "").strip(),
            "source_file":    str(r["source_file"] or "").strip(),
            "check_number":   re.sub(r"[^0-9]", "", str(r["check_number"] or "")),
        })

    written, refreshed = 0, set()
    for safe, sc_rows in by_safe.items():
        folder = processed_root / safe
        if not folder.exists():
            continue
        try:
            ip._write_sidecar(folder, sc_rows)
            written += 1
            refreshed.add(folder.resolve())
        except OSError as exc:
            print(f"  [!] Could not write {ip.SIDECAR_NAME} for '{safe}' ({exc})")
    for sc in processed_root.glob("*/" + ip.SIDECAR_NAME):     # empty out stale sidecars
        if sc.parent.resolve() not in refreshed:
            try:
                ip._write_sidecar(sc.parent, [])
            except OSError:
                pass
    return written


# --------------------------------------------------------------------------- xlsx snapshot

def export_to_xlsx(path: Path = None) -> Path:
    """
    Optional on-demand snapshot of the invoice log to a styled workbook (All Invoices + one tab
    per property), for anyone who still wants an Excel copy. Not part of the pipeline - the DB
    is the source of truth. Reuses the processor's header/style constants.
    """
    import openpyxl
    path = Path(path or (config.DATA_DIR / "invoices_export.xlsx"))
    invoices = list_invoices()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # master sheet
    master = wb.create_sheet(ip.MASTER_SHEET)
    ip._style_header_row(master, ip.MASTER_HEADER)
    ip._set_col_widths(master, ip.MASTER_WIDTHS)
    # per-property sheets built as we go
    for inv in reversed(invoices):                 # oldest first, like the log grew
        amount_cell = inv["amount"] if inv["amount"] is not None else (inv["amount_text"] or "")
        master_vals = [
            inv["status"], inv["vendor_name"], inv["invoice_number"], inv["unit"],
            inv["invoice_date"], inv["due_date"], amount_cell, inv["description"],
            inv["line_items"], inv["property"], inv["source_file"], inv["date_processed"],
            ip.YARDI_CHECK if inv["entered_in_yardi"] else "", inv["stored_file"],
            inv["reconciled"], inv["check_number"], inv["carried_forward"],
        ]
        ip._append_styled_row(master, ip.MASTER_HEADER, master_vals,
                              duplicate=(inv["status"] == "DUPLICATE"))
        if inv["status"] != "DUPLICATE":
            tab = ip._sanitize_sheet_name(
                ip.PROPERTY_REVIEW_TAB if inv["needs_review"] else (inv["property"] or "Unknown Property"))
            ws = wb[tab] if tab in wb.sheetnames else None
            if ws is None:
                ws = wb.create_sheet(tab)
                ip._style_header_row(ws, ip.PROPERTY_HEADER)
                ip._set_col_widths(ws, ip.PROPERTY_WIDTHS)
            ip._append_styled_row(ws, ip.PROPERTY_HEADER, [
                inv["vendor_name"], inv["invoice_number"], inv["unit"], inv["invoice_date"],
                inv["due_date"], amount_cell, inv["description"], inv["line_items"],
                inv["source_file"], inv["date_processed"],
                ip.YARDI_CHECK if inv["entered_in_yardi"] else "", inv["stored_file"],
                inv["reconciled"], inv["check_number"], inv["carried_forward"],
            ])
    wb.save(path)
    return path
