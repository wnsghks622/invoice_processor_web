# -*- coding: utf-8 -*-
"""
migrate_to_db.py - one-time import of the current live data into this app's SQLite DB.

    python migrate_to_db.py          (imports; refuses if the DB already has invoices)
    python migrate_to_db.py --force  (wipes and re-imports)

Reads the ORIGINAL invoice_processor/ files READ-ONLY (invoices.xlsx, properties.csv,
vendors.csv) and never writes them - the original folder stays a pristine fallback.

Invoices are imported as the UNION of the master 'All Invoices' sheet AND every property tab,
because an audit found ~21 invoices that live only on property tabs (the tabs and master had
drifted from hand-edits). A tab row is treated as the SAME invoice as a master row when they
share a loose identity (vendor + invoice# + unit); only tab rows whose identity appears nowhere
in the master are imported as genuine tab-only extras. A reconciliation report is printed at the
end for review before cut-over.
"""
from __future__ import annotations

import argparse
import csv
import sys

import openpyxl

import config
from core import db
from core import processor as ip

RESERVED = {"Summary", "All Invoices"}
REVIEW_TAB = "Needs Review"


def norm(s):
    return ip._normalize_key(str(s or ""))


def loose_identity(vendor, invoice_no, unit, stored_file, amount_raw, invoice_date):
    """Identity used to tell whether a tab row is the same invoice as a master row. A real
    invoice number makes vendor+invoice#+unit unique; without one, fall back to fields that
    separate distinct bills."""
    ninv = norm(invoice_no)
    if ninv:
        return ("n", norm(vendor), ninv, norm(unit))
    amt = ip._parse_amount(amount_raw)
    return ("e", norm(vendor), norm(unit), norm(stored_file),
            f"{amt:.2f}" if amt is not None else "", norm(invoice_date))


def amount_fields(raw):
    """(amount_real_or_None, amount_text) - number when parseable, else keep the raw text."""
    amt = ip._parse_amount(raw)
    if amt is not None:
        return amt, ""
    return None, (str(raw).strip() if raw not in (None, "") else "")


def yardi_flag(v):
    return 1 if str(v or "").strip() else 0


def seed_properties() -> int:
    if not config.ORIGINAL_PROPERTIES_CSV.exists():
        print(f"  [!] {config.ORIGINAL_PROPERTIES_CSV} not found - no properties seeded.")
        return 0
    n = 0
    with config.ORIGINAL_PROPERTIES_CSV.open(newline="", encoding="utf-8-sig") as f:
        for i, row in enumerate(csv.DictReader(f)):
            name = (row.get("canonical_name") or "").strip()
            if not name:
                continue
            db.add_property(name, (row.get("property_code") or "").strip(),
                            (row.get("aliases") or "").strip(), sort_order=i)
            n += 1
    return n


def seed_vendors() -> int:
    if not config.ORIGINAL_VENDORS_CSV.exists():
        print(f"  [!] {config.ORIGINAL_VENDORS_CSV} not found - no vendors seeded.")
        return 0
    n = 0
    with config.ORIGINAL_VENDORS_CSV.open(newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            short = (row.get("short_name") or "").strip()
            if not short:
                continue
            db.add_vendor(short, (row.get("aliases") or "").strip())
            n += 1
    return n


def _row_dict(header, row):
    return {header[i]: row[i] if i < len(row) else None for i in range(len(header))}


def import_invoices(canonical_names: set):
    """Returns (master_count, tab_only_rows, master_loose_set). Inserts as it goes."""
    wb = openpyxl.load_workbook(str(config.ORIGINAL_XLSX), read_only=True, data_only=True)
    canon_by_sheet = {ip._sanitize_sheet_name(c): c for c in canonical_names}

    # ---- master ----
    master = wb["All Invoices"]
    mh = [c.value for c in next(master.iter_rows(min_row=1, max_row=1))]
    master_loose = set()
    master_count = 0
    for row in master.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        d = _row_dict(mh, row)
        vendor = d.get("Vendor Name")
        stored = d.get("Stored File")
        if not (vendor or stored):
            continue
        status = str(d.get("Status") or "OK").strip() or "OK"
        prop = str(d.get("Property") or "").strip()
        amt, amt_text = amount_fields(d.get("Amount"))
        needs_review = 1 if (status != "DUPLICATE" and prop and prop not in canonical_names) else 0
        db.insert_invoice({
            "status": status, "vendor_name": str(vendor or "").strip(),
            "invoice_number": str(d.get("Invoice #") or "").strip(),
            "unit": str(d.get("Unit") or "").strip(),
            "invoice_date": str(d.get("Invoice Date") or "").strip(),
            "due_date": str(d.get("Due Date") or "").strip(),
            "amount": amt, "amount_text": amt_text,
            "description": str(d.get("Description") or "").strip(),
            "line_items": str(d.get("Line Items (review)") or "").strip(),
            "property": prop, "source_file": str(d.get("Source File") or "").strip(),
            "date_processed": str(d.get("Date Processed") or "").strip(),
            "entered_in_yardi": yardi_flag(d.get("Entered in Yardi")),
            "stored_file": str(stored or "").strip(),
            "reconciled": str(d.get("Reconciled") or "").strip(),
            "check_number": str(d.get("Check #") or "").strip(),
            "carried_forward": str(d.get("Carried Forward") or "").strip(),
            "needs_review": needs_review, "origin": "master",
        })
        master_count += 1
        master_loose.add(loose_identity(vendor, d.get("Invoice #"), d.get("Unit"),
                                        stored, d.get("Amount"), d.get("Invoice Date")))

    # ---- property tabs (union) ----
    tab_only = []
    seen_tab = set()
    for sheet in wb.sheetnames:
        if sheet in RESERVED:
            continue
        ws = wb[sheet]
        h = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        is_review = (sheet == REVIEW_TAB)
        canonical = "" if is_review else canon_by_sheet.get(sheet, sheet)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            d = _row_dict(h, row)
            vendor = d.get("Vendor Name")
            stored = d.get("Stored File")
            if not (vendor or stored):
                continue
            ident = loose_identity(vendor, d.get("Invoice #"), d.get("Unit"),
                                   stored, d.get("Amount"), d.get("Invoice Date"))
            if ident in master_loose or ident in seen_tab:
                continue                       # same invoice as a master (or earlier tab) row
            seen_tab.add(ident)
            amt, amt_text = amount_fields(d.get("Amount"))
            db.insert_invoice({
                "status": "OK", "vendor_name": str(vendor or "").strip(),
                "invoice_number": str(d.get("Invoice #") or "").strip(),
                "unit": str(d.get("Unit") or "").strip(),
                "invoice_date": str(d.get("Invoice Date") or "").strip(),
                "due_date": str(d.get("Due Date") or "").strip(),
                "amount": amt, "amount_text": amt_text,
                "description": str(d.get("Description") or "").strip(),
                "line_items": str(d.get("Line Items (review)") or "").strip(),
                "property": canonical,
                "source_file": str(d.get("Source File") or "").strip(),
                "date_processed": str(d.get("Date Processed") or "").strip(),
                "entered_in_yardi": yardi_flag(d.get("Entered in Yardi")),
                "stored_file": str(stored or "").strip(),
                "reconciled": str(d.get("Reconciled") or "").strip(),
                "check_number": str(d.get("Check #") or "").strip(),
                "carried_forward": str(d.get("Carried Forward") or "").strip(),
                "needs_review": 1 if is_review else 0,
                "origin": f"tab:{sheet}",
            })
            tab_only.append({"tab": sheet, "vendor": str(vendor or "").strip(),
                             "invoice": str(d.get("Invoice #") or "").strip(),
                             "amount": amt, "stored": str(stored or "").strip()})
    wb.close()
    return master_count, tab_only, master_loose


def report():
    invs = db.list_invoices()
    total = len(invs)
    dup = sum(1 for i in invs if i["status"] == "DUPLICATE")
    recon = sum(1 for i in invs if (i["reconciled"] or "").strip())
    yardi = sum(1 for i in invs if i["entered_in_yardi"])
    review = sum(1 for i in invs if i["needs_review"])
    amt_sum = sum(i["amount"] for i in invs if i["amount"] is not None)
    no_amt = sum(1 for i in invs if i["amount"] is None)
    print("\n" + "=" * 64)
    print("  MIGRATION REPORT")
    print("=" * 64)
    print(f"  invoices in DB ....... {total}   (OK {total - dup} / DUPLICATE {dup})")
    print(f"  reconciled ........... {recon}")
    print(f"  entered in Yardi ..... {yardi}")
    print(f"  needs review ......... {review}")
    print(f"  amount total ......... ${amt_sum:,.2f}   ({no_amt} row(s) with no numeric amount)")


def main():
    ap = argparse.ArgumentParser(description="Import the live xlsx/CSV data into SQLite")
    ap.add_argument("--force", action="store_true", help="wipe existing DB rows and re-import")
    args = ap.parse_args()

    if not config.ORIGINAL_XLSX.exists():
        print(f"[ERROR] Original spreadsheet not found:\n  {config.ORIGINAL_XLSX}")
        sys.exit(1)

    db.init()
    if not db.is_empty():
        if not args.force:
            print("[ABORT] The database already has invoices. Re-run with --force to wipe and re-import.")
            sys.exit(1)
        with db._connect() as conn:
            conn.execute("DELETE FROM invoices")
            conn.execute("DELETE FROM properties")
            conn.execute("DELETE FROM vendors")
        print("[*] --force: cleared existing rows.")

    print("[*] Seeding properties and vendors...")
    nprop = seed_properties()
    nvend = seed_vendors()
    canonical = {p["name"] for p in db.all_properties()}
    print(f"    {nprop} properties, {nvend} vendors.")

    print("[*] Importing invoices (master + property tabs, de-duplicated)...")
    master_count, tab_only, _ = import_invoices(canonical)
    print(f"    {master_count} master rows imported; {len(tab_only)} tab-only invoice(s) added.")

    if tab_only:
        print("\n  Tab-only invoices (present on a property tab but NOT in the master log):")
        for t in tab_only:
            amt = f"${t['amount']:,.2f}" if t["amount"] is not None else "(no amt)"
            print(f"    [{t['tab']:>24}] {t['vendor'][:26]:26} #{t['invoice'][:14]:14} "
                  f"{amt:>12}  {t['stored']}")

    report()
    print("\n[*] Done. Review the tab-only list above before relying on the DB.")


if __name__ == "__main__":
    main()
